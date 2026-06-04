"""
ReceiptVault Excel Generator
Produces a 3-sheet accountant-ready workbook:
  Sheet 1 — Inventory Purchases
  Sheet 2 — Business Expenses  
  Sheet 3 — Income Summary
"""

import io
from datetime import datetime
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Colors ───────────────────────────────────────────────────────────────────
GREEN_DARK  = "1a6b3a"
GREEN_LIGHT = "e8f4ec"
GREEN_MID   = "2d8c52"
GRAY_HEADER = "f4f4f3"
GRAY_ROW    = "fafafa"
WHITE       = "ffffff"
BLACK       = "000000"

INVENTORY_CATEGORIES = [
    "Inventory", "Meals & Entertainment", "Travel",
    "Vehicle & Fuel", "Equipment", "Other"
]

EXPENSE_CATEGORIES = [
    "Telephone", "Heat / Utilities", "Insurance",
    "Business Rent", "Employee Wages",
    "Office Supplies", "Advertising",
    "Software & Subscriptions", "Other Expenses"
]

def _thick():
    return Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def _bottom():
    return Border(bottom=Side(style="medium"))

def _header_font(size=11):
    return Font(name="Arial", bold=True, color=WHITE, size=size)

def _body_font(bold=False, size=10):
    return Font(name="Arial", bold=bold, size=size, color=BLACK)

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _currency(ws, cell_ref):
    ws[cell_ref].number_format = '$#,##0.00'

def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def _title_row(ws, text, row, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    c.fill = _fill(GREEN_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28

def _col_header(ws, row, cols):
    for i, label in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = _header_font(10)
        c.fill = _fill(GREEN_MID)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _thick()
    ws.row_dimensions[row].height = 20


def build_excel(
    business_name: str,
    period_label: str,
    receipts: List[Dict[str, Any]],
    gross_income: float = 0.0,
) -> bytes:
    wb = Workbook()

    # ── Separate receipts by type ─────────────────────────────────────────────
    inventory = [r for r in receipts if (r.get("category") or "Other") in INVENTORY_CATEGORIES]
    expenses  = [r for r in receipts if (r.get("category") or "Other") not in INVENTORY_CATEGORIES
                 or (r.get("category") or "Other") == "Other"]

    # group expenses by category
    expense_map: Dict[str, List] = {}
    for r in receipts:
        cat = r.get("category") or "Other"
        if cat not in INVENTORY_CATEGORIES or cat == "Other":
            expense_map.setdefault(cat, []).append(r)

    # ── SHEET 1: Inventory Purchases ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Inventory Purchases"

    _title_row(ws1, f"{business_name.upper()} — {period_label.upper()} INVENTORY PURCHASES", 1, 4)
    _col_header(ws1, 2, ["Vendor / Supplier", "Invoice / Ref", "Date", "Amount"])
    _set_col_widths(ws1, {"A": 30, "B": 20, "C": 14, "D": 14})

    row = 3
    inv_rows = [r for r in receipts if (r.get("category") or "") == "Inventory"]
    for i, r in enumerate(inv_rows):
        merchant = r.get("merchant") or r.get("originalName") or r.get("original_name") or "Unknown"
        ref = r.get("notes") or "—"
        date_raw = r.get("receiptDate") or r.get("receipt_date") or r.get("uploadedAt") or r.get("uploaded_at") or ""
        try:
            date_str = datetime.fromisoformat(str(date_raw)[:10]).strftime("%m/%d/%Y")
        except Exception:
            date_str = str(date_raw)[:10] if date_raw else "—"
        amt = r.get("amount")
        bg = WHITE if i % 2 == 0 else GRAY_ROW

        ws1.cell(row=row, column=1, value=merchant).font = _body_font()
        ws1.cell(row=row, column=2, value=ref).font = _body_font()
        ws1.cell(row=row, column=3, value=date_str).font = _body_font()
        amt_cell = ws1.cell(row=row, column=4, value=float(amt) if amt else 0)
        amt_cell.font = _body_font()
        amt_cell.number_format = '$#,##0.00'

        for col in range(1, 5):
            c = ws1.cell(row=row, column=col)
            c.fill = _fill(bg)
            c.alignment = Alignment(vertical="center")
        row += 1

    # Total row
    if inv_rows:
        total_formula = f"=SUM(D3:D{row-1})"
    else:
        total_formula = "=0"

    ws1.cell(row=row, column=1, value="TOTAL INVENTORY PURCHASES").font = Font(name="Arial", bold=True, size=10, color=WHITE)
    ws1.cell(row=row, column=1).fill = _fill(GREEN_DARK)
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    t = ws1.cell(row=row, column=4, value=total_formula)
    t.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    t.fill = _fill(GREEN_DARK)
    t.number_format = '$#,##0.00'
    ws1.row_dimensions[row].height = 22

    # ── SHEET 2: Business Expenses ────────────────────────────────────────────
    ws2 = wb.create_sheet("Business Expenses")
    _title_row(ws2, f"{business_name.upper()} — {period_label.upper()} BUSINESS EXPENSES", 1, 3)
    _col_header(ws2, 2, ["Expense Category", "Vendor / Notes", "Amount"])
    _set_col_widths(ws2, {"A": 28, "B": 30, "C": 14})

    row2 = 3
    non_inv = [r for r in receipts if (r.get("category") or "Other") not in ["Inventory"]]
    for i, r in enumerate(non_inv):
        cat = r.get("category") or "Other Expenses"
        merchant = r.get("merchant") or r.get("originalName") or r.get("original_name") or "Unknown"
        amt = r.get("amount")
        bg = WHITE if i % 2 == 0 else GRAY_ROW

        ws2.cell(row=row2, column=1, value=cat).font = _body_font()
        ws2.cell(row=row2, column=2, value=merchant).font = _body_font()
        a = ws2.cell(row=row2, column=3, value=float(amt) if amt else 0)
        a.font = _body_font()
        a.number_format = '$#,##0.00'
        for col in range(1, 4):
            ws2.cell(row=row2, column=col).fill = _fill(bg)
            ws2.cell(row=row2, column=col).alignment = Alignment(vertical="center")
        row2 += 1

    # Total
    exp_total = f"=SUM(C3:C{row2-1})" if non_inv else "=0"
    ws2.cell(row=row2, column=1, value="TOTAL BUSINESS EXPENSES").font = Font(name="Arial", bold=True, size=10, color=WHITE)
    ws2.cell(row=row2, column=1).fill = _fill(GREEN_DARK)
    ws2.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=2)
    t2 = ws2.cell(row=row2, column=3, value=exp_total)
    t2.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    t2.fill = _fill(GREEN_DARK)
    t2.number_format = '$#,##0.00'

    # ── SHEET 3: Income Summary ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Income Summary")
    _title_row(ws3, f"{business_name.upper()} — {period_label.upper()} INCOME SUMMARY", 1, 3)
    _col_header(ws3, 2, ["LINE ITEM", "NOTES", "AMOUNT"])
    _set_col_widths(ws3, {"A": 42, "B": 28, "C": 16})

    now = datetime.now()
    summary_rows = [
        ("1.  Gross Business Income", period_label, gross_income, False),
        ("2.  Inventory Purchases (Cost of Goods)", "See 'Inventory Purchases' sheet", f"='Inventory Purchases'!D{len(inv_rows)+3}", False),
        ("3.  Gross Income After Inventory  (Line 1 – Line 2)", "", "=C3-C4", True),
    ]

    r3 = 3
    for label, notes, val, bold in summary_rows:
        ws3.cell(row=r3, column=1, value=label).font = _body_font(bold=bold)
        ws3.cell(row=r3, column=2, value=notes).font = _body_font()
        c = ws3.cell(row=r3, column=3, value=val)
        c.font = _body_font(bold=bold)
        c.number_format = '$#,##0.00'
        bg = GREEN_LIGHT if bold else WHITE
        for col in range(1, 4):
            ws3.cell(row=r3, column=col).fill = _fill(bg)
            ws3.cell(row=r3, column=col).alignment = Alignment(vertical="center")
        r3 += 1

    # Expenses header
    ws3.cell(row=r3, column=1, value="BUSINESS EXPENSES (DEDUCTIONS)").font = Font(name="Arial", bold=True, size=10, color=WHITE)
    ws3.cell(row=r3, column=1).fill = _fill(GREEN_MID)
    ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=3)
    r3 += 1

    exp_start = r3
    for i, r in enumerate(non_inv):
        cat = r.get("category") or "Other"
        merchant = r.get("merchant") or ""
        amt = r.get("amount")
        bg = WHITE if i % 2 == 0 else GRAY_ROW
        ws3.cell(row=r3, column=1, value=f"    {cat}").font = _body_font()
        ws3.cell(row=r3, column=2, value=merchant).font = _body_font()
        c = ws3.cell(row=r3, column=3, value=float(amt) if amt else 0)
        c.font = _body_font()
        c.number_format = '$#,##0.00'
        for col in range(1, 4):
            ws3.cell(row=r3, column=col).fill = _fill(bg)
        r3 += 1

    # Total expenses + net income
    total_exp_formula = f"=SUM(C{exp_start}:C{r3-1})" if non_inv else "=0"

    ws3.cell(row=r3, column=1, value="    TOTAL Business Expenses").font = _body_font(bold=True)
    ws3.cell(row=r3, column=1).fill = _fill(GRAY_HEADER)
    ws3.cell(row=r3, column=2).fill = _fill(GRAY_HEADER)
    t3 = ws3.cell(row=r3, column=3, value=total_exp_formula)
    t3.font = _body_font(bold=True)
    t3.fill = _fill(GRAY_HEADER)
    t3.number_format = '$#,##0.00'
    r3 += 1

    # NET INCOME
    ws3.cell(row=r3, column=1, value="NET INCOME  (Line 3 – Total Expenses)").font = Font(name="Arial", bold=True, size=11, color=WHITE)
    ws3.cell(row=r3, column=1).fill = _fill(GREEN_DARK)
    ws3.cell(row=r3, column=2).fill = _fill(GREEN_DARK)
    ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=2)
    net = ws3.cell(row=r3, column=3, value=f"=C5-C{r3-1}")
    net.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    net.fill = _fill(GREEN_DARK)
    net.number_format = '$#,##0.00'
    ws3.row_dimensions[r3].height = 26

    # ── Freeze panes and finalize ─────────────────────────────────────────────
    for ws in [ws1, ws2, ws3]:
        ws.freeze_panes = "A3"
        ws.sheet_view.showGridLines = True

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
